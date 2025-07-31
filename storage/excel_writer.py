import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def save_to_excel(leads, filename="leads.xlsx"):
    df = pd.DataFrame(leads)
    df.to_excel(filename, index=False)

    try:
        wb = load_workbook(filename)
        ws = wb.active

        header_fill = PatternFill(start_color="BFD7EA", end_color="BFD7EA", fill_type="solid")
        for col_num, cell in enumerate(ws[1], 1):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = header_fill
            max_length = max(len(str(cell.value or "")), 12)
            ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

        wb.save(filename)
        print(f"✅ Excel file saved: {filename}")

    except Exception as e:
        print(f"⚠️ Excel formatting error: {e}")



def save_to_csv(leads, filename="leads.csv"):
    df = pd.DataFrame(leads)
    df.to_csv(filename, index=False)
    print(f"✅ CSV file saved: {filename}")